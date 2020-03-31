import os
import cv2
import csv
import random
import shutil
import numpy as np

from glob import glob
from openpyxl import Workbook, load_workbook


def evaluate(ca_path='D:/data/geek/close_contours/ca.csv',
             output={
                 'hor01_004_k_A.A0019.png': {},
                 'hor01_018_021_k_A.A0005.png': {'1356_1862': '1364_1868', '200_10': '10_0'}
             },
             report_path='report.xlsx'):
    """

    Args:
        ca_path (str): file path to ca (.csv)
        output (dict): in the form {'img_name_with_extension': {'row1_col1': 'row2_col2', 'row1_col1': 'row2_col2'}
        report_path (str): file path to report (.xlsx). If report_path is not None, write report to file.

    Returns:
        img_report (dict): the result of each image
        row_report (dict): the result of each point
        summary_report (dict): the overall result

    """
    # Read from CA.
    with open(ca_path) as csv_file:
        ca = list(csv.reader(csv_file, delimiter=','))
    csv_file.close()

    # Turn CA list into a dict for easier processing.
    ca_dict = {}
    for i, row in enumerate(ca[1:]):
        img_name, _, _, _, n_contours, row1, col1, row2, col2 = row
        # Turn points to strings to make a dict.
        coor1 = '%s_%s' % (row1, col1)
        coor2 = '%s_%s' % (row2, col2)
        row_index = i + 1
        if img_name in ca_dict:
            ca_dict[img_name]['coor_info'].append((coor1, coor2, row_index))
        else:
            ca_dict[img_name] = {'n_contours': int(n_contours), 'coor_info': [(coor1, coor2, row_index)]}

    img_report = {}
    row_report = {}
    total_img = len(output)
    total_correct_img = total_points = total_fn = total_fp = total_tp = 0
    for img_name, output_coords in output.items():
        # For each new image, set True Positive to 0.
        tp = 0
        img_report[img_name] = {
            'precision': '-',
            'recall': '-',
            'n_contours': 0,
            'n_output': 0,
            'correct': 0,
            'overall': 0
        }

        # Initally, set False Negative equal to number of contours in CA and decrement by one as the process goes on.
        img_report[img_name]['n_contours'] = fn = ca_dict[img_name]['n_contours']
        # Similarly, set False Positive equal to number of output pairs.
        img_report[img_name]['n_output'] = fp = len(output_coords)
        # Increment the number of total segments.
        total_points += fn
        # Get the inverted dictionary.
        inverted_output = {y: x for x, y in output_coords.items()}

        # Process each segment in each image in CA.
        for info in ca_dict[img_name]['coor_info']:
            coor1, coor2, row_index = info
            result = 0  # if the pair is correct, result = 1. Otherwise, 0.
            if output_coords.get(coor1) == coor2 or inverted_output.get(coor1) == coor2:
                tp += 1
                fn -= 1
                fp -= 1
                result = 1
            row_report[row_index] = result

        # Only calculate P and R if n_contours is not 0.
        if ca_dict[img_name]['n_contours'] != 0:
            precision = tp * 1.0 / (tp + fp)
            recall = tp * 1.0 / (tp + fn)
            img_report[img_name]['precision'] = '%.2f' % precision
            img_report[img_name]['recall'] = '%.2f' % recall

        img_report[img_name]['correct'] = tp
        if img_report[img_name]['correct'] == img_report[img_name]['n_output']:
            img_report[img_name]['overall'] = 1
            total_correct_img += 1

        total_tp += tp
        total_fp += fp
        total_fn += fn

    total_precision = total_recall = percent_correct = 0.00
    if total_points != 0:
        total_precision = total_tp * 1.0 / (total_tp + total_fp)
        total_recall = total_tp * 1.0 / (total_tp + total_fn)
        percent_correct = total_tp * 1.0 / total_points
    percent_correct_img = total_correct_img * 1.0 / total_img
    summary_report = {
        'overall_precision_point': '%.2f' % total_precision,
        'overall_recall_point': '%.2f' % total_recall,
        'percent_correct_point': '%.2f' % percent_correct,
        'total_img': total_img,
        'total_correct_img': total_correct_img,
        'per_correct_img': '%.2f' % percent_correct_img
    }

    if report_path:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'point_report'
        titles = ['File_Name', 'Cut_Name', 'Cut_Reference', 'Image_reference', 'Number_of_Open_Contours',
                  'Start_Row_(Y)', 'Start_Col_(X)', 'End_Row_(Y)', 'End_Col_(X)', 'Result']
        ws1.append(titles)
        with open(ca_path) as csv_file:
            ca = list(csv.reader(csv_file, delimiter=','))
        csv_file.close()
        for i, ca_row in enumerate(ca[1:]):
            new_row = ca_row + [row_report.get(i + 1)]
            ws1.append(new_row)

        ws2 = wb.create_sheet('img_report')
        ws2.append(('Image_Name', 'n_contours', 'n_output', 'Total_Correct', 'Precision', 'Recall'))
        for img_name, report_info in img_report.items():
            ws2.append(
                (
                    img_name,
                    report_info['n_contours'],
                    report_info['n_output'],
                    report_info['correct'],
                    report_info['precision'],
                    report_info['recall']
                )
            )

        ws3 = wb.create_sheet('summary')
        for k, v in summary_report.items():
            ws3.append((k, v))
        wb.save(report_path)

    return img_report, row_report, summary_report


def pick_reference_sketch(sketch_dir=None, sketch_paths=None):
    """Picks the sketch with the most components as the reference sketch.

    Args:
        sketch_dir (str): path to the folder that contains all sketches in one cut
        sketch_paths (list): list of all sketch paths in sorted order

    Returns:
        str: the name of the reference image
        str: the path of the reference image
        int: the index of the reference image in the list of the sketch paths
    """
    from src.component_utils.component_wrapper import ComponentWrapper
    from natsort import natsorted
    from src.utils import imread

    if sketch_dir:
        sketch_paths = natsorted(glob('%s/*' % sketch_dir))

    # Find the sketch with the most components as the reference.
    max = -1
    for index, sketch_path in enumerate(sketch_paths):
        img = imread(sketch_path, grayscale=True)

        components_properties = ComponentWrapper().process(img, is_gray=True)
        components = components_properties[1]

        if len(components) > max:
            max = len(components)
            ref_sketch_path = sketch_path
            ref_index = index
    ref_img_name = os.path.basename(ref_sketch_path)
    return ref_img_name, ref_sketch_path, ref_index


def choose_img_randomly(
        paths=None,
        keep=1,
        remove=10,
        input_dir='D:/data/geek/original/hor01_sketches_cat',
        output_dir='D:/data/geek/hor01_sketches_cat'
):
    paths = glob('%s/*' % input_dir)
    for path in paths:
        print(path)
        name = os.path.basename(path)
        output_path = '%s/%s' % (output_dir, name)
        # if random.randint(1, keep + remove) > keep:
        #     os.remove(path)
        # if random.randint(1, keep + remove) <= keep:
        img = 255 - cv2.imread(path)
        if np.sum(img) < 620000:
            shutil.move(path, output_path)


def thin_img(img=None):
    from src.preprocessing_utils.close_object_v2 import do_thin
    output_dir = 'D:/data/geek/hor01_sketches_cat'
    img_paths = glob('D:/data/geek/HOR01_trace_all/*')
    for img_path in img_paths:
        print(img_path)
        img_name = os.path.basename(img_path)[:-4]
        img = cv2.imread(img_path, 0)
        out = 255 - do_thin(img)
        cv2.imwrite('%s/%s.png' % (output_dir, img_name), out)


def create_open_contours(start_mode=True):
    from natsort import natsorted
    input_dir = 'D:/data/geek/close_contours/original'
    output_dir = 'D:/data/geek/close_contours/input'
    debug_dir = 'D:/data/geek/close_contours/debug'
    excel_path = 'D:/data/geek/close_contours/ca.xlsx'
    paths = natsorted(glob('%s/*' % input_dir))

    if start_mode:
        #####################
        # CREATE EXCEL FILE #
        #####################
        # Create excel file with the title row first and save.
        wb = Workbook()
        ws = wb.active

        titles = ['File_Name', 'Cut_Name', 'Cut_Reference', 'Image_reference', 'Number_of_Open_Contours',
                  'Start_Row_(Y)', 'Start_Col_(X)', 'End_Row_(Y)', 'End_Col_(X)']
        ws.append(titles)
        wb.save(excel_path)

    check = 0
    for k, path in enumerate(paths[check:]):
        print('%d: %s' % (k + check, path))
        # Load the excel file
        wb = load_workbook(excel_path)
        ws = wb.active

        ##################
        # FIND REFERENCE #
        ##################
        img_name = os.path.basename(path)
        name_list = img_name.split('.')
        cut_name = name_list[0]

        # Find the first reference for the cut.
        cut_dir = 'D:/data/geek/original/hor01_sketches_cat'
        sketch_paths = natsorted(glob('%s/%s*' % (cut_dir, cut_name)))
        first_ref_img_name, first_ref_img_path, _ = pick_reference_sketch(sketch_paths=sketch_paths)

        # Find the reference of the current sketch.
        _img_name = name_list[-2]  # name without extension
        img_index = int(_img_name[-4:]) - 1
        if img_index == 0:
            ref_img_name = os.path.basename(sketch_paths[1])
        else:
            ref_img_name = os.path.basename(sketch_paths[img_index - 1])
        excel_row = [img_name, cut_name, first_ref_img_name, ref_img_name]

        # Read the image
        org_img = 255 - cv2.imread(path)
        debug_img = org_img.copy()  # to visualize the endpoints of open contours for debugging
        out_img = org_img.copy()  # output image

        # Only open contours in 2/3 of the total images.
        if random.randint(1, 3) == 3:
            out_excel_row = excel_row + [0, -1, -1, -1, -1]
            ws.append(out_excel_row)

        else:
            #################
            # OPEN CONTOURS #
            #################
            gray = 255 - cv2.imread(path, 0)
            gray_img = gray.copy()  # need a gray image to find contours' locations more easily when using np.argwhere

            n_row, n_col = np.shape(gray_img)
            locations = np.argwhere(gray_img)  # list of coordinates of all pixels of all contours
            # Choose how many segments to erase.
            i = n_seg = random.randint(1, 30)
            while n_seg > 0:
                # Choose a starting point.
                start = random.randint(0, len(locations) - 1)
                cur_row, cur_col = locations[start]
                out_excel_row = excel_row + [i, cur_row, cur_col]
                # Chose how many pixels to erase.
                j = n_pix = random.randint(2, 30)

                # Erase the point.
                gray_img[cur_row, cur_col] = 0
                # Draw a red-filled circle around the starting point:
                debug_img = cv2.circle(debug_img, (cur_col, cur_row), 3, (255, 255, 0), -1)
                while 0 < cur_row < n_row - 1 and 0 < cur_col < n_col - 1 and n_pix > 0:
                    n_pix -= 1
                    # Zoom in on the neighbor pixels
                    temp = gray_img[cur_row-1:cur_row+2, cur_col-1:cur_col+2].copy()
                    # Get a list of the neighbors' coordinates
                    neighbors = np.argwhere(temp)
                    n_neighbors = len(neighbors)
                    # If there is no neighbor, continue.
                    if n_neighbors == 0:
                        continue
                    # Pick a neighbor randomly.
                    row_increment, col_increment = neighbors[random.randint(0, n_neighbors - 1)]
                    cur_row += row_increment - 1
                    cur_col += col_increment - 1

                    # Erase the neighbor's pixel.
                    out_img[cur_row, cur_col, :] = (0, 0, 0)
                    debug_img[cur_row, cur_col, :] = (0, 0, 255)
                    gray_img[cur_row, cur_col] = 0

                # If the number of erased pixels is less than 2, continue.
                # This means at least 2 pixels must have been erased (one of them being the end point).
                if j - n_pix < 2:
                    continue

                # Reduce the number of segments:
                n_seg -= 1
                # Restore the end point for the output image.
                out_img[cur_row, cur_col, :] = (255, 255, 255)
                out_excel_row += [cur_row, cur_col]
                # Draw a red-filled circle around the end point in the debug_img.
                debug_img = cv2.circle(debug_img, (cur_col, cur_row), 3, (255, 255, 0), -1)
                # Erase the end point in the debug_img.
                debug_img[cur_row, cur_col, :] = (0, 0, 0)

                # Update locations.
                locations = np.argwhere(gray_img)
                # Update excel sheet.
                ws.append(out_excel_row)

        cv2.imwrite('%s/%s' % (debug_dir, img_name), 255 - debug_img)
        cv2.imwrite('%s/%s' % (output_dir, img_name), 255 - out_img)
        wb.save(excel_path)

