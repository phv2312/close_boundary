import csv
from algo.rule1.cat_utils import *


def evaluate(ca_path, output, report_path=None, **kwargs):
    """Evaluates closing contours.

    Args:
        ca_path (str): file path to ca (.csv)
        output (dict): in the form {'img_name_with_extension': {'row1_col1': 'row2_col2', 'row1_col1': 'row2_col2', ...}
        report_path (str): file path to report (.xlsx). If report_path is not None, write report to file.

    Returns:
        img_report (dict): the result of each image
        row_report (dict): the result of each point
        summary_report (dict): the overall result

    """
    # Read from CA.
    ca_dict = process_CA_for_evaluation(ca_path)
    img_report = {}
    row_report = {}
    total_img = len(output)
    total_correct_img = total_points = total_fn = total_fp = total_tp = 0
    for img_name, output_coords in output.items():
        # For each new image, set True Positive to 0.
        tp = 0
        img_report[img_name] = {
            'precision': '-',
            'recall': '-',
            'n_contours': 0,
            'n_output': 0,
            'correct': 0,
            'overall': 0
        }

        # Initally, set False Negative equal to number of contours in CA and decrement by one as the process goes on.
        img_report[img_name]['n_contours'] = fn = ca_dict[img_name]['n_contours']
        # Similarly, set False Positive equal to number of output pairs.
        img_report[img_name]['n_output'] = fp = len(output_coords)
        # Increment the number of total segments.
        total_points += fn
        # Get the inverted dictionary.
        inverted_output = {y: x for x, y in output_coords.items()}

        # Process each segment in each image in CA.
        for info in ca_dict[img_name]['coor_info']:
            coor1, coor2, row_index = info
            result = 0  # if the pair is correct, result = 1. Otherwise, 0.
            if output_coords.get(coor1) == coor2 or inverted_output.get(coor1) == coor2:
                tp += 1
                fn -= 1
                fp -= 1
                result = 1
            row_report[row_index] = result

        # Only calculate P and R if n_contours is not 0.
        if ca_dict[img_name]['n_contours'] != 0 and (tp + fp) != 0:
            precision = tp * 1.0 / (tp + fp)
        if ca_dict[img_name]['n_contours'] != 0 and (tp + fn) != 0:
            recall = tp * 1.0 / (tp + fn)
            img_report[img_name]['precision'] = '%.2f' % precision
            img_report[img_name]['recall'] = '%.2f' % recall

        img_report[img_name]['correct'] = tp
        if img_report[img_name]['correct'] == img_report[img_name]['n_output'] == img_report[img_name]['n_contours']:
            img_report[img_name]['overall'] = 1
            total_correct_img += 1

        total_tp += tp
        total_fp += fp
        total_fn += fn

    total_precision = total_recall = percent_correct = 0.00
    if total_points != 0:
        total_precision = total_tp * 1.0 / (total_tp + total_fp)
        total_recall = total_tp * 1.0 / (total_tp + total_fn)
        percent_correct = total_tp * 1.0 / total_points
    percent_correct_img = total_correct_img * 1.0 / total_img
    _summary_report = {
        'overall_precision_point': '%.2f' % total_precision,
        'overall_recall_point': '%.2f' % total_recall,
        'percent_correct_point': '%.2f' % percent_correct,
        '-': '-',
        'total_img': total_img,
        'total_correct_img': total_correct_img,
        'percent_correct_img': '%.2f' % percent_correct_img
    }
    summary_report = {**kwargs, **_summary_report}

    if report_path:
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'point_report'
        titles = ['File_Name', 'Cut_Name', 'Cut_Reference', 'Image_reference', 'Number_of_Open_Contours',
                  'Start_Row_(Y)', 'Start_Col_(X)', 'End_Row_(Y)', 'End_Col_(X)', 'Result']
        ws1.append(titles)
        with open(ca_path) as csv_file:
            ca = list(csv.reader(csv_file, delimiter=','))
        csv_file.close()
        for i, ca_row in enumerate(ca[1:]):
            new_row = ca_row + [row_report.get(i + 1)]
            ws1.append(new_row)

        ws2 = wb.create_sheet('img_report')
        ws2.append(('Image_Name', 'n_contours', 'n_output', 'Total_Correct', 'Precision', 'Recall', 'Overall'))
        for img_name, report_info in img_report.items():
            ws2.append(
                (
                    img_name,
                    report_info['n_contours'],
                    report_info['n_output'],
                    report_info['correct'],
                    report_info['precision'],
                    report_info['recall'],
                    report_info['overall']
                )
            )

        ws3 = wb.create_sheet('summary')
        for k, v in summary_report.items():
            ws3.append((k, v))
        wb.save(report_path)

    return img_report, row_report, summary_report


if __name__ == '__main__':
    import os
    import numpy as np
    from PIL import Image
    from glob import glob
    from datetime import datetime
    from algo.rule1.closing import ClosingModel
    from algo.rule1.utils import convert_output_format

    # TODO:
    dev_mode = True  # whether we're processing actual clients' sketches
    visualize_mode = True  # whether we wanna visualize debug images

    closing_model = ClosingModel(max_pair_distance=10, max_traveled_pixel=10, keypoint_to_boundary_distance=7)
    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y_%H-%M-%S")

    # TODO:
    paths = glob('D:/data/geek/close_contours/input/*')
    ca_path = 'D:/data/geek/close_contours/ca.csv'
    ref_colored_img_dir = 'D:/data/geek/HOR01_deta_all'
    debug_img_dir = 'D:/data/geek/close_contours/debug'

    # TODO:
    output_dir = 'D:/output/geek/close_contours/%s' % dt_string
    os.makedirs(output_dir, exist_ok=True)
    report_path = '%s/report_%s.xlsx' % (output_dir, dt_string)

    result = {}
    preprocess_ca_dict = process_CA_for_preprocessing(ca_path)
    for i, p in enumerate(paths):
        print(i, p)
        # Load the image as a gray image.
        sketch_tgt_im = np.array(Image.open(p).convert('L'))

        if dev_mode:
            # Preprocess.
            sketch_fn = os.path.basename(p)
            ref_colored_img_fn = preprocess_ca_dict[sketch_fn]
            ref_colored_img_path = '%s/%s' % (ref_colored_img_dir, ref_colored_img_fn)
            ref_colored_img = np.asarray(Image.open(ref_colored_img_path))
            sketch, interval = preprocess_imgs(sketch_tgt_im, ref_colored_img)
            sketch = sketch[0]
            up_bound, _ = interval
        else:
            sketch = sketch_tgt_im
            up_bound = 0

        # Close.
        pair_points, max_traveled_pixels, max_pair_distance, keypoint_to_boundary_distance = \
            closing_model.process(sketch)
        result[sketch_fn] = convert_output_format(pair_points, up_bound)

        # Visualize the debug image.
        if visualize_mode:
            debug_img_path = '%s/%s' % (debug_img_dir, sketch_fn)
            visualize_debug_img(debug_img_path, result[sketch_fn], mode='evaluation', output_dir=output_dir)

    # Export report.xlsx.
    kwargs = {
        'max_traveled_pixels': max_traveled_pixels,
        'max_pair_distance': max_pair_distance,
        'keypoint_to_boundary_distance': keypoint_to_boundary_distance,
        '-': '-'
    }
    evaluate(ca_path, result, report_path, **kwargs)
